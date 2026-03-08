"""
create_status_report.py — Generate an engagement status report Excel workbook.

Produces a multi-sheet .xlsx file with:
  Sheet 1: Portfolio Summary   — all clients with per-service progress and key metrics
  Sheet 2+: Per-Client Detail  — one sheet per client with service progress, PBC, queries, deadlines

Usage (as module):
    from create_status_report import create_status_report
    create_status_report(engagements, output_path)

Usage (standalone):
    python create_status_report.py \
        --pm-dir "C:\\Users\\khjan\\Downloads\\Pilot - Project Manager Main Agent" \
        --output "C:\\...\\Status_Report.xlsx"
"""

import os
import re
import sys
import argparse
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Style constants — matching create_consolidated_pbc.py conventions
# ---------------------------------------------------------------------------

HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="Arial", size=10)
DATA_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
DATA_CENTER = Alignment(horizontal="center", vertical="top")

TITLE_FONT = Font(name="Arial", size=14, bold=True, color="1F4E79")
SUBTITLE_FONT = Font(name="Arial", size=11, bold=False, color="666666")
SECTION_FONT = Font(name="Arial", size=11, bold=True, color="1F4E79")
TOTALS_FONT = Font(name="Arial", size=10, bold=True)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Progress colour fills
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_GREY = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# Status colour fills (for PBC items on detail sheets)
FILL_RECEIVED = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_PENDING = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_OUTSTANDING = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

FILL_ALT_ROW = PatternFill(start_color="F5F8FC", end_color="F5F8FC", fill_type="solid")


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def _get_progress_fill(pct: int) -> PatternFill:
    """Return colour fill based on progress percentage."""
    if pct > 75:
        return FILL_GREEN
    elif pct >= 25:
        return FILL_YELLOW
    else:
        return FILL_RED


def _get_status_fill(status: str) -> PatternFill:
    """Return fill colour for a PBC status string."""
    s = status.strip().lower()
    if s in ("received", "derived"):
        return FILL_RECEIVED
    elif s == "pending":
        return FILL_PENDING
    elif s in ("not applicable", "not_applicable", "n/a"):
        return FILL_GREY
    else:
        return FILL_OUTSTANDING


def _apply_header_row(ws, row_num: int, columns: list):
    """Apply header styling to a row given column definitions [(name, width), ...]."""
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width


def _write_data_row(ws, row_num: int, values: list, fill=None, bold=False):
    """Write a list of values as a data row with standard styling."""
    font = TOTALS_FONT if bold else DATA_FONT
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font = font
        cell.alignment = DATA_ALIGNMENT
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def _safe_sheet_name(name: str) -> str:
    """Sanitise a string for use as an Excel sheet name.

    - Max 31 characters
    - Remove characters not allowed in sheet names: [ ] : * ? / \\
    - Strip leading/trailing whitespace
    """
    clean = re.sub(r'[\[\]:*?/\\]', '', name).strip()
    # Shorten common suffixes
    clean = clean.replace("Sdn Bhd", "").replace("Sdn. Bhd.", "").strip()
    if len(clean) > 31:
        clean = clean[:28] + "..."
    return clean if clean else "Client"


def _format_stage(stage: str) -> str:
    """Convert an internal stage code to a human-readable label."""
    labels = {
        "not_started": "Not Started",
        "planning": "Planning",
        "fieldwork": "Fieldwork",
        "completion": "Completion",
        "done": "Completed",
        "in_progress": "In Progress",
        "pbc_pending": "PBC Pending",
        "review": "Review",
        "data_received": "Data Received",
        "generated": "FS Generated",
        "pending_pbc": "PBC Pending",
        "completed": "Completed",
    }
    return labels.get(stage, stage.replace("_", " ").title())


def _find_next_deadline(engagement: dict) -> str:
    """Find the earliest upcoming deadline from an engagement dict."""
    today = datetime.now().strftime("%Y-%m-%d")
    candidates = []

    # Per-service deadlines
    for svc_key, svc_data in engagement.get("services", {}).items():
        if svc_data.get("enabled") and svc_data.get("deadline"):
            dl = svc_data["deadline"]
            if dl >= today:
                candidates.append(dl)

    # Top-level deadlines
    for dl in engagement.get("deadlines", []):
        date_val = dl.get("date", dl.get("deadline", ""))
        if date_val and date_val >= today:
            candidates.append(date_val)

    if candidates:
        candidates.sort()
        return candidates[0]
    return ""


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_portfolio_summary(ws, engagements: list):
    """Build the Portfolio Summary sheet."""
    ws.title = "Portfolio Summary"

    # Title block
    row = 1
    cell = ws.cell(row=row, column=1, value="Engagement Portfolio Status Report")
    cell.font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)

    row = 2
    cell = ws.cell(row=row, column=1, value=f"Generated: {datetime.now().strftime('%d %B %Y, %I:%M %p')}")
    cell.font = SUBTITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)

    # Column definitions for the summary table
    columns = [
        ("Client", 30),
        ("Audit Status", 16),
        ("Audit %", 10),
        ("Tax Status", 16),
        ("Tax %", 10),
        ("Compilation Status", 18),
        ("Compilation %", 14),
        ("PBC Outstanding", 16),
        ("Queries Open", 14),
        ("Next Deadline", 16),
    ]

    row = 4
    _apply_header_row(ws, row, columns)

    # Data rows
    total_pbc_outstanding = 0
    total_queries_open = 0
    data_start_row = row + 1

    row = data_start_row
    for eng in engagements:
        client = eng.get("client", {})
        services = eng.get("services", {})
        probe = eng.get("_probe", {})

        client_name = client.get("display_name") or client.get("legal_name", "Unknown")

        # Audit
        audit_cfg = services.get("audit", {})
        if audit_cfg.get("enabled"):
            audit_probe = probe.get("audit", {})
            audit_stage = _format_stage(audit_probe.get("stage", audit_cfg.get("status", "not_started")))
            audit_pct = audit_probe.get("estimated_progress", audit_cfg.get("progress_pct", 0))
        else:
            audit_stage = "N/A"
            audit_pct = ""

        # Tax
        tax_cfg = services.get("tax", {})
        if tax_cfg.get("enabled"):
            tax_probe = probe.get("tax", {})
            tax_stage = _format_stage(tax_probe.get("stage", tax_cfg.get("status", "not_started")))
            tax_pct = tax_probe.get("estimated_progress", tax_cfg.get("progress_pct", 0))
        else:
            tax_stage = "N/A"
            tax_pct = ""

        # Compilation
        comp_cfg = services.get("compilation", {})
        if comp_cfg.get("enabled"):
            comp_probe = probe.get("compilation", {})
            comp_stage = _format_stage(comp_probe.get("stage", comp_cfg.get("status", "not_started")))
            comp_pct = comp_probe.get("estimated_progress", comp_cfg.get("progress_pct", 0))
        else:
            comp_stage = "N/A"
            comp_pct = ""

        # PBC and queries
        pbc_summary = eng.get("pbc_summary", {})
        pbc_outstanding = pbc_summary.get("outstanding", 0)
        queries_summary = eng.get("queries_summary", {})
        queries_open = queries_summary.get("open", 0)

        total_pbc_outstanding += pbc_outstanding
        total_queries_open += queries_open

        # Next deadline
        next_deadline = _find_next_deadline(eng)

        values = [
            client_name,
            audit_stage,
            audit_pct if audit_pct != "" else "—",
            tax_stage,
            tax_pct if tax_pct != "" else "—",
            comp_stage,
            comp_pct if comp_pct != "" else "—",
            pbc_outstanding,
            queries_open,
            next_deadline if next_deadline else "—",
        ]
        _write_data_row(ws, row, values)

        # Apply progress colour coding to percentage cells
        if isinstance(audit_pct, (int, float)):
            ws.cell(row=row, column=3).fill = _get_progress_fill(audit_pct)
            ws.cell(row=row, column=3).alignment = DATA_CENTER
        if isinstance(tax_pct, (int, float)):
            ws.cell(row=row, column=5).fill = _get_progress_fill(tax_pct)
            ws.cell(row=row, column=5).alignment = DATA_CENTER
        if isinstance(comp_pct, (int, float)):
            ws.cell(row=row, column=7).fill = _get_progress_fill(comp_pct)
            ws.cell(row=row, column=7).alignment = DATA_CENTER

        # Colour code PBC outstanding (red if > 0)
        if pbc_outstanding > 0:
            ws.cell(row=row, column=8).fill = FILL_RED
        # Colour code queries open (yellow if > 0)
        if queries_open > 0:
            ws.cell(row=row, column=9).fill = FILL_YELLOW

        row += 1

    # Totals row
    if engagements:
        totals = [
            f"TOTAL ({len(engagements)} clients)",
            "", "", "", "", "", "",
            total_pbc_outstanding,
            total_queries_open,
            "",
        ]
        _write_data_row(ws, row, totals, bold=True, fill=FILL_GREY)

    # Freeze panes
    ws.freeze_panes = ws.cell(row=5, column=1)

    # Auto-filter
    if engagements:
        ws.auto_filter.ref = f"A4:J{row - 1}"


def _build_client_detail_sheet(ws, engagement: dict):
    """Build a per-client detail sheet."""
    client = engagement.get("client", {})
    services = engagement.get("services", {})
    probe = engagement.get("_probe", {})

    client_name = client.get("legal_name", "Unknown")
    display_name = client.get("display_name", client_name)

    ws.title = _safe_sheet_name(display_name or client_name)

    # --- Client Header ---
    row = 1
    cell = ws.cell(row=row, column=1, value=display_name or client_name)
    cell.font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

    row = 2
    fye = client.get("fye_date", "")
    reg_no = client.get("registration_no", "")
    details = []
    if reg_no:
        details.append(f"Reg: {reg_no}")
    if fye:
        details.append(f"FYE: {fye}")
    if client.get("reporting_framework"):
        details.append(f"Framework: {client['reporting_framework']}")
    cell = ws.cell(row=row, column=1, value="  |  ".join(details))
    cell.font = SUBTITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

    row = 3
    cell = ws.cell(row=row, column=1, value=f"Report Date: {datetime.now().strftime('%d %B %Y')}")
    cell.font = SUBTITLE_FONT

    # --- Service Progress Table ---
    row = 5
    cell = ws.cell(row=row, column=1, value="Service Progress")
    cell.font = SECTION_FONT

    row = 6
    svc_cols = [("Service", 20), ("Status", 16), ("Progress %", 12), ("Deadline", 16), ("Notes", 40)]
    _apply_header_row(ws, row, svc_cols)

    row = 7
    for svc_key in ("audit", "tax", "compilation"):
        svc_cfg = services.get(svc_key, {})
        if not svc_cfg.get("enabled"):
            values = [svc_key.capitalize(), "Not Enabled", "—", "—", "—"]
            _write_data_row(ws, row, values, fill=FILL_GREY)
        else:
            svc_probe = probe.get(svc_key, {})
            stage = _format_stage(svc_probe.get("stage", svc_cfg.get("status", "not_started")))
            pct = svc_probe.get("estimated_progress", svc_cfg.get("progress_pct", 0))
            deadline = svc_cfg.get("deadline", "")
            notes = svc_cfg.get("notes", "")

            values = [svc_key.capitalize(), stage, pct, deadline if deadline else "—", notes if notes else ""]
            progress_fill = _get_progress_fill(pct)
            _write_data_row(ws, row, values)
            ws.cell(row=row, column=3).fill = progress_fill
            ws.cell(row=row, column=3).alignment = DATA_CENTER
        row += 1

    # --- Outstanding PBC Items ---
    row += 1
    cell = ws.cell(row=row, column=1, value="Outstanding PBC Items")
    cell.font = SECTION_FONT
    row += 1

    pbc_summary = engagement.get("pbc_summary", {})
    total_pbc = pbc_summary.get("total_items", 0)
    received = pbc_summary.get("received", 0)
    outstanding = pbc_summary.get("outstanding", 0)

    if total_pbc > 0:
        pbc_info_cols = [("Metric", 20), ("Count", 12)]
        _apply_header_row(ws, row, pbc_info_cols)
        row += 1
        _write_data_row(ws, row, ["Total Items", total_pbc])
        row += 1
        _write_data_row(ws, row, ["Received", received], fill=FILL_GREEN if received > 0 else None)
        row += 1
        _write_data_row(ws, row, ["Outstanding", outstanding], fill=FILL_RED if outstanding > 0 else None)
        row += 1
        _write_data_row(ws, row, ["Not Applicable", pbc_summary.get("not_applicable", 0)])
        row += 1
    else:
        cell = ws.cell(row=row, column=1, value="No PBC data available. Run /consolidated-pbc first.")
        cell.font = Font(name="Arial", size=10, italic=True, color="999999")
        row += 1

    # --- Open Queries ---
    row += 1
    cell = ws.cell(row=row, column=1, value="Queries Summary")
    cell.font = SECTION_FONT
    row += 1

    queries_summary = engagement.get("queries_summary", {})
    total_queries = queries_summary.get("total", 0)
    open_queries = queries_summary.get("open", 0)
    resolved_queries = queries_summary.get("resolved", 0)

    if total_queries > 0:
        q_cols = [("Metric", 20), ("Count", 12)]
        _apply_header_row(ws, row, q_cols)
        row += 1
        _write_data_row(ws, row, ["Total Queries", total_queries])
        row += 1
        _write_data_row(ws, row, ["Open", open_queries], fill=FILL_YELLOW if open_queries > 0 else None)
        row += 1
        _write_data_row(ws, row, ["Resolved", resolved_queries], fill=FILL_GREEN if resolved_queries > 0 else None)
        row += 1
    else:
        cell = ws.cell(row=row, column=1, value="No queries recorded. Run /consolidated-queries first.")
        cell.font = Font(name="Arial", size=10, italic=True, color="999999")
        row += 1

    # --- Deadlines ---
    row += 1
    cell = ws.cell(row=row, column=1, value="Deadlines")
    cell.font = SECTION_FONT
    row += 1

    deadlines = engagement.get("deadlines", [])

    # Also collect per-service deadlines
    all_deadlines = []
    for svc_key in ("audit", "tax", "compilation"):
        svc_cfg = services.get(svc_key, {})
        if svc_cfg.get("enabled") and svc_cfg.get("deadline"):
            all_deadlines.append({
                "description": f"{svc_key.capitalize()} deadline",
                "date": svc_cfg["deadline"],
                "service": svc_key,
            })
    for dl in deadlines:
        all_deadlines.append({
            "description": dl.get("description", ""),
            "date": dl.get("date", dl.get("deadline", "")),
            "service": dl.get("service", ""),
        })

    if all_deadlines:
        dl_cols = [("Description", 35), ("Date", 16), ("Service", 14)]
        _apply_header_row(ws, row, dl_cols)
        row += 1

        all_deadlines.sort(key=lambda d: d.get("date", ""))
        today = datetime.now().strftime("%Y-%m-%d")

        for dl in all_deadlines:
            values = [
                dl.get("description", ""),
                dl.get("date", ""),
                dl.get("service", "").capitalize(),
            ]
            # Highlight overdue deadlines in red
            dl_date = dl.get("date", "")
            if dl_date and dl_date < today:
                _write_data_row(ws, row, values, fill=FILL_RED)
            else:
                _write_data_row(ws, row, values)
            row += 1
    else:
        cell = ws.cell(row=row, column=1, value="No deadlines set.")
        cell.font = Font(name="Arial", size=10, italic=True, color="999999")

    # Set column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 16


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def create_status_report(engagements: list, output_path: str) -> str:
    """Generate the engagement status report Excel workbook.

    Args:
        engagements: List of engagement dicts. Each dict should contain the
                     full engagement.json structure plus an optional '_probe'
                     key with the output of StatusReader.probe_all_services().
        output_path: Absolute path for the output .xlsx file.

    Returns:
        Absolute path of the saved workbook.
    """
    wb = Workbook()

    # Sheet 1: Portfolio Summary (the default sheet)
    ws_summary = wb.active
    _build_portfolio_summary(ws_summary, engagements)

    # Sheet 2+: Per-Client Detail
    # Track sheet names to avoid duplicates
    used_names = {"Portfolio Summary"}
    for eng in engagements:
        ws = wb.create_sheet()
        _build_client_detail_sheet(ws, eng)

        # Ensure unique sheet name
        base_name = ws.title
        if base_name in used_names:
            counter = 2
            while f"{base_name} ({counter})" in used_names:
                counter += 1
            ws.title = f"{base_name} ({counter})"
        used_names.add(ws.title)

    # Ensure output directory exists
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.isdir(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    wb.save(output_path)
    print(f"Status report saved: {output_path}")
    return os.path.abspath(output_path)


# ---------------------------------------------------------------------------
# Standalone CLI
# ---------------------------------------------------------------------------

def main():
    """Entry point for standalone execution."""
    parser = argparse.ArgumentParser(
        description="Generate an engagement status report Excel workbook."
    )
    parser.add_argument(
        "--pm-dir",
        required=True,
        help="Path to the Project Manager agent root directory.",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output .xlsx path. Defaults to <pm-dir>/Clients/Portfolio_Status_Report.xlsx",
    )
    parser.add_argument(
        "--client",
        default=None,
        help="Generate report for a specific client only (by folder name).",
    )
    args = parser.parse_args()

    # Add PM dir to path so we can import tools
    sys.path.insert(0, args.pm_dir)
    from tools.pm_engine import PMEngine
    from tools.status_reader import StatusReader

    engine = PMEngine(clients_dir=os.path.join(args.pm_dir, "Clients"))
    reader = StatusReader()

    print("Scanning engagements...")
    all_summaries = engine.list_engagements()

    if args.client:
        all_summaries = [s for s in all_summaries if s["folder"] == args.client]
        if not all_summaries:
            print(f"ERROR: No engagement found for client '{args.client}'")
            sys.exit(1)

    engagements = []
    for summary in all_summaries:
        folder = summary["folder"]
        print(f"  Loading: {folder}")
        try:
            eng_data = engine.load_engagement(folder)
        except FileNotFoundError:
            print(f"    WARNING: engagement.json not found, skipping.")
            continue

        # Probe sub-agent folders for live status
        print(f"    Probing sub-agent folders...")
        probe_results = reader.probe_all_services(eng_data)
        eng_data["_probe"] = probe_results
        engagements.append(eng_data)

    if not engagements:
        print("No engagements found. Nothing to report.")
        sys.exit(0)

    print(f"\nGenerating report for {len(engagements)} engagement(s)...")

    # Determine output path
    if args.output:
        output_path = args.output
    else:
        if args.client:
            output_dir = os.path.join(engine.clients_dir, args.client, "output")
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, f"{args.client}_Status_Report.xlsx")
        else:
            output_path = os.path.join(engine.clients_dir, "Portfolio_Status_Report.xlsx")

    create_status_report(engagements, output_path)
    print("Done.")


if __name__ == "__main__":
    main()
