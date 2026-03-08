"""
create_consolidated_pbc.py — Generate a consolidated PBC Excel workbook.

Produces a multi-sheet .xlsx file with:
  Sheet 1: Summary        — totals, by-service breakdown, by-category breakdown
  Sheet 2: All Items      — full PBC list with colour-coded status rows
  Sheet 3: Audit PBC      — filtered to audit items only
  Sheet 4: Tax PBC        — filtered to tax items only
  Sheet 5: Compilation PBC — filtered to compilation items only

Usage (as module):
    from create_consolidated_pbc import create_consolidated_pbc
    create_consolidated_pbc(client_name, items, stats, output_path)

Usage (standalone):
    python create_consolidated_pbc.py --client "ABC Trading Sdn Bhd" \
        --pm-dir "C:\\Users\\khjan\\Downloads\\Pilot - Project Manager Main Agent"
"""

import os
import sys
import argparse
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Style constants — shared across all sheets
# ---------------------------------------------------------------------------

HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="Arial", size=10)
DATA_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

TITLE_FONT = Font(name="Arial", size=14, bold=True, color="1F4E79")
SUBTITLE_FONT = Font(name="Arial", size=11, bold=False, color="666666")
SECTION_FONT = Font(name="Arial", size=11, bold=True, color="1F4E79")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Status colour fills
FILL_RECEIVED = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_PENDING = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_OUTSTANDING = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_NA = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# Light alternating row fill (for summary tables)
FILL_ALT_ROW = PatternFill(start_color="F5F8FC", end_color="F5F8FC", fill_type="solid")

# Column definitions for the PBC item tables
PBC_COLUMNS = [
    ("Ref", 8),
    ("Document", 35),
    ("Category", 22),
    ("Needed For", 22),
    ("Status", 14),
    ("Date Received", 16),
    ("Priority", 10),
    ("Remarks", 40),
]


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def _get_status_fill(status: str) -> PatternFill:
    """Return the appropriate fill colour for a status string."""
    s = status.strip().lower()
    if s in ("received", "derived"):
        return FILL_RECEIVED
    elif s == "pending":
        return FILL_PENDING
    elif s in ("not applicable", "not_applicable", "n/a"):
        return FILL_NA
    else:
        return FILL_OUTSTANDING


def _apply_header_row(ws, row_num: int, columns: list):
    """Apply header styling to a row given column definitions [(name, width), ...]."""
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width


def _write_data_row(ws, row_num: int, values: list, fill=None):
    """Write a list of values as a data row with standard styling."""
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font = DATA_FONT
        cell.alignment = DATA_ALIGNMENT
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def _write_pbc_items(ws, items: list, start_row: int) -> int:
    """Write PBC items to a worksheet starting at start_row. Returns next empty row."""
    # Header
    _apply_header_row(ws, start_row, PBC_COLUMNS)

    row = start_row + 1
    for item in items:
        needed_for = ", ".join(
            s.capitalize() for s in item.get("needed_for", [])
        )
        values = [
            item.get("ref", ""),
            item.get("document", ""),
            item.get("category", ""),
            needed_for,
            item.get("status", ""),
            item.get("date_received", ""),
            (item.get("priority", "medium") or "medium").capitalize(),
            item.get("remarks", ""),
        ]
        status_fill = _get_status_fill(item.get("status", ""))
        _write_data_row(ws, row, values, fill=status_fill)
        row += 1

    # Freeze panes at header row
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    # Auto-filter
    if len(items) > 0:
        last_col_letter = get_column_letter(len(PBC_COLUMNS))
        ws.auto_filter.ref = f"A{start_row}:{last_col_letter}{row - 1}"

    return row


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_summary_sheet(ws, client_name: str, items: list, stats: dict):
    """Build the Summary sheet with totals and breakdowns."""
    ws.title = "Summary"

    # Title
    row = 1
    cell = ws.cell(row=row, column=1, value=f"Consolidated PBC Checklist")
    cell.font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

    row = 2
    cell = ws.cell(row=row, column=1, value=client_name)
    cell.font = Font(name="Arial", size=12, bold=True, color="333333")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

    row = 3
    cell = ws.cell(row=row, column=1, value=f"Generated: {datetime.now().strftime('%d %B %Y, %I:%M %p')}")
    cell.font = SUBTITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

    # --- Overall Summary Table ---
    row = 5
    cell = ws.cell(row=row, column=1, value="Overall Summary")
    cell.font = SECTION_FONT

    row = 6
    summary_cols = [("Metric", 22), ("Count", 12)]
    _apply_header_row(ws, row, summary_cols)

    summary_data = [
        ("Total Documents", stats.get("total", 0)),
        ("Received", stats.get("received", 0)),
        ("Outstanding", stats.get("outstanding", 0)),
        ("Not Applicable", stats.get("not_applicable", 0)),
    ]

    row = 7
    for metric, count in summary_data:
        _write_data_row(ws, row, [metric, count])
        row += 1

    # --- By-Service Breakdown ---
    row += 1
    cell = ws.cell(row=row, column=1, value="By Service")
    cell.font = SECTION_FONT
    row += 1

    svc_cols = [("Service", 18), ("Total", 10), ("Received", 12), ("Outstanding", 14)]
    _apply_header_row(ws, row, svc_cols)
    row += 1

    by_service = stats.get("by_service", {})
    for svc_key in ("audit", "tax", "compilation"):
        svc_data = by_service.get(svc_key, {"total": 0, "received": 0, "outstanding": 0})
        values = [
            svc_key.capitalize(),
            svc_data.get("total", 0),
            svc_data.get("received", 0),
            svc_data.get("outstanding", 0),
        ]
        alt_fill = FILL_ALT_ROW if svc_key == "tax" else None
        _write_data_row(ws, row, values, fill=alt_fill)
        row += 1

    # --- By-Category Breakdown ---
    row += 1
    cell = ws.cell(row=row, column=1, value="By Category")
    cell.font = SECTION_FONT
    row += 1

    cat_cols = [("Category", 28), ("Total", 10), ("Received", 12), ("Outstanding", 14)]
    _apply_header_row(ws, row, cat_cols)
    row += 1

    by_category = stats.get("by_category", {})
    for idx, (cat_name, cat_data) in enumerate(sorted(by_category.items())):
        values = [
            cat_name,
            cat_data.get("total", 0),
            cat_data.get("received", 0),
            cat_data.get("outstanding", 0),
        ]
        alt_fill = FILL_ALT_ROW if idx % 2 == 1 else None
        _write_data_row(ws, row, values, fill=alt_fill)
        row += 1

    # Set column widths for summary sheet
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16


def _build_all_items_sheet(ws, items: list):
    """Build the 'All Items' sheet with the full PBC list."""
    ws.title = "All Items"
    _write_pbc_items(ws, items, start_row=1)


def _build_filtered_sheet(ws, title: str, items: list, service_key: str):
    """Build a filtered sheet showing only items for a specific service."""
    ws.title = title
    filtered = [
        item for item in items
        if service_key in item.get("needed_for", [])
    ]

    if filtered:
        _write_pbc_items(ws, filtered, start_row=1)
    else:
        ws.cell(row=1, column=1, value=f"No {title.lower()} items found.").font = SUBTITLE_FONT
        ws.column_dimensions["A"].width = 40


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def create_consolidated_pbc(client_name: str, items: list, stats: dict, output_path: str) -> str:
    """Generate the consolidated PBC Excel workbook.

    Args:
        client_name: Display name of the client (used in title).
        items:       List of merged PBC item dicts from PBCConsolidator.consolidate().
        stats:       Summary stats dict from PBCConsolidator.generate_summary_stats().
        output_path: Absolute path for the output .xlsx file.

    Returns:
        Absolute path of the saved workbook.
    """
    wb = Workbook()

    # Sheet 1: Summary (the default sheet created with Workbook())
    ws_summary = wb.active
    _build_summary_sheet(ws_summary, client_name, items, stats)

    # Sheet 2: All Items
    ws_all = wb.create_sheet()
    _build_all_items_sheet(ws_all, items)

    # Sheet 3: Audit PBC
    ws_audit = wb.create_sheet()
    _build_filtered_sheet(ws_audit, "Audit PBC", items, "audit")

    # Sheet 4: Tax PBC
    ws_tax = wb.create_sheet()
    _build_filtered_sheet(ws_tax, "Tax PBC", items, "tax")

    # Sheet 5: Compilation PBC
    ws_comp = wb.create_sheet()
    _build_filtered_sheet(ws_comp, "Compilation PBC", items, "compilation")

    # Ensure output directory exists
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.isdir(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    wb.save(output_path)
    print(f"PBC workbook saved: {output_path}")
    return os.path.abspath(output_path)


# ---------------------------------------------------------------------------
# Standalone CLI
# ---------------------------------------------------------------------------

def main():
    """Entry point for standalone execution."""
    parser = argparse.ArgumentParser(
        description="Generate a consolidated PBC Excel workbook."
    )
    parser.add_argument(
        "--client",
        required=True,
        help="Client legal name (must match a folder in Clients/).",
    )
    parser.add_argument(
        "--pm-dir",
        required=True,
        help="Path to the Project Manager agent root directory.",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output .xlsx path. Defaults to Clients/<client>/output/<client>_PBC_Consolidated.xlsx",
    )
    args = parser.parse_args()

    # Add PM dir to path so we can import tools
    sys.path.insert(0, args.pm_dir)
    from tools.pm_engine import PMEngine
    from tools.pbc_consolidator import PBCConsolidator

    engine = PMEngine(clients_dir=os.path.join(args.pm_dir, "Clients"))

    print(f"Loading engagement for: {args.client}")
    engagement = engine.load_engagement(args.client)

    print("Consolidating PBC items...")
    consolidator = PBCConsolidator(engagement)
    items, stats = consolidator.consolidate()

    print(f"  Total: {stats['total']}, Received: {stats['received']}, "
          f"Outstanding: {stats['outstanding']}, N/A: {stats['not_applicable']}")

    # Determine output path
    if args.output:
        output_path = args.output
    else:
        output_dir = os.path.join(engine.clients_dir, args.client, "output")
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f"{args.client}_PBC_Consolidated.xlsx")

    create_consolidated_pbc(args.client, items, stats, output_path)
    print("Done.")


if __name__ == "__main__":
    main()
